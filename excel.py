from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill
from smGenerator import smgenerator

def bonjeanGenerator(sample, path):


    depthOfVessel = 20

    # EXTRACTING VALUES FROM WORKSHEET
    wb = load_workbook(sample)
    ws = wb.active
    stations = []
    waterlines = []
    offsetTable = []

    for i, row in enumerate(ws['A'][1:]):
        stations.append(row.value)
        offsetTable.append([])
        for column in ws[i+2][1:]:
            offsetTable[i].append(column.value)

    for column in ws[1][1:]:
        waterlines.append(column.value)

    wb.close()

    wtlnSpacing = depthOfVessel / waterlines[-1]
    sm = smgenerator(waterlines)


    # FOR CUMULATIVE AREA
    smArea = []
    for n, station in enumerate(offsetTable):
        i = 0
        smArea.append([])
        for multipliers in sm:
            area = 0
            for each in multipliers:
                area += each * station[i]
                i += 1
            smArea[n].append(area)
            i -= 1

    cumulativeArea = []
    for i, stationAreas in enumerate(smArea):
        cumulativeArea.append([])
        sum = 0
        for each in stationAreas:
            sum += (1/3) * wtlnSpacing * each
            cumulativeArea[i].append(sum)




    # FOR CUMULATIVE MOMENT 
    lever = []
    for i in waterlines:
        lever.append(i * wtlnSpacing)

    smMoment = []
    for n, station in enumerate(offsetTable):
        i = 0
        smMoment.append([])
        for multipliers in sm:
            moment = 0
            for each in multipliers:
                moment += each * station[i] * lever[i]
                # print(each, lever[i], station[i])
                i += 1
            smMoment[n].append(moment)
            i -= 1

    cumulativeMoment = []
    for i, stationMoments in enumerate(smMoment):
        cumulativeMoment.append([])
        sum = 0
        for each in stationMoments:
            sum += (1/3) * wtlnSpacing * each
            cumulativeMoment[i].append(sum)

    # FOR Z 
    Z = []
    for i in range(len(smArea)):
        Z.append([])
        for j in range(len(smArea[i])):
            if smArea[i][j] == 0:
                Z[i].append(0)
            else:
                Z[i].append(smMoment[i][j] / smArea[i][j])


    # CREATING NEW WORKBOOK 
    wb = Workbook()
    template = wb.active

    header = ["STATION",0]
    title = ["WL",	"d",	"ΣfAs",	"As",	"ΣfM𝘇",	"M𝘇",	"Z"]
    template.append(header)
    template.append(title)

    displayWtlns = []
    n = 0
    for i in sm:
        n += (len(i) - 1)
        displayWtlns.append(waterlines[n])


    for i, value in enumerate(displayWtlns):
        template.cell(row=i+3, column=1, value=value)
        template.cell(row=i+3, column=2, value=value * wtlnSpacing)
        for j in range(2,8):
            template.cell(row=i+3, column=j).fill = PatternFill("solid", fgColor="0099CCFF")


    for i in range(1,3):
        for cell in template[i]:
            cell.font = Font(bold=True, color="00FFFFFF")
            cell.fill = PatternFill("solid", fgColor="00003366")

    for cell in template['A']:
        cell.font = Font(bold=True, color="00FFFFFF")
        cell.fill = PatternFill("solid", fgColor="00003366")

    for n, station in enumerate(stations):
        worksheet = wb.copy_worksheet(template)
        worksheet.title = f"STATION {station}"
        worksheet.cell(row=1, column=2, value=station)
        for e, each in enumerate([smArea, cumulativeArea, smMoment, cumulativeMoment, Z]):
            for i, area in enumerate(each[n]):
                worksheet.cell(row=i+3, column=e+3, value=area)

    del wb["Sheet"]

    wb.save(path + "/bonjeanSheet.xlsx")